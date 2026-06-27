"""Excel export service for generating formatted reports."""

from datetime import date
from io import BytesIO
from decimal import Decimal

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.attendance import AttendanceRecord, OvertimeRecord
from app.models.bonus import Bonus, BonusType, Reimbursement, ReimbursementType, THRRecord
from app.models.employee import Department, Employee
from app.models.kasbon import KasbonRequest
from app.models.payroll import PayrollRun, Payslip


class ExcelExportService:
    """Service for exporting data to formatted Excel files."""

    @staticmethod
    def _freeze_header(ws):
        """Freeze the header row."""
        ws.freeze_panes = "A2"

    @staticmethod
    def _auto_width(ws):
        """Auto-fit column widths based on content."""
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except (TypeError, AttributeError):
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = min(adjusted_width, 50)

    @staticmethod
    def _apply_currency_format(ws, cols: list[int]):
        """Apply currency number format to specified columns (1-indexed), skipping header."""
        for row in ws.iter_rows(min_row=2):
            for col_idx in cols:
                cell = row[col_idx - 1]
                cell.number_format = "#,##0"

    @staticmethod
    def _decimal_to_float(val) -> float:
        """Convert Decimal or numeric to float for Excel."""
        if val is None:
            return 0.0
        if isinstance(val, Decimal):
            return float(val)
        return float(val)

    @staticmethod
    def export_payslips(payroll_run_id: int, db: Session) -> bytes:
        """Export payslips for a payroll run as formatted Excel."""
        payslips = (
            db.query(Payslip)
            .filter(Payslip.payroll_run_id == payroll_run_id)
            .all()
        )

        # Load employee info
        employee_ids = [p.employee_id for p in payslips]
        employees = {
            e.id: e
            for e in db.query(Employee).filter(Employee.id.in_(employee_ids)).all()
        } if employee_ids else {}

        # Load department info
        dept_ids = [e.department_id for e in employees.values() if e.department_id]
        departments = {
            d.id: d.name
            for d in db.query(Department).filter(Department.id.in_(dept_ids)).all()
        } if dept_ids else {}

        rows = []
        for ps in payslips:
            emp = employees.get(ps.employee_id)
            dept_name = departments.get(emp.department_id, "") if emp and emp.department_id else ""
            rows.append({
                "Kode Karyawan": emp.employee_code if emp else "",
                "Nama Lengkap": emp.full_name if emp else "",
                "Departemen": dept_name,
                "Gaji Pokok": ExcelExportService._decimal_to_float(ps.basic_salary),
                "Total Tunjangan": ExcelExportService._decimal_to_float(ps.total_allowances),
                "Lembur": ExcelExportService._decimal_to_float(ps.overtime_amount),
                "Bonus": ExcelExportService._decimal_to_float(ps.bonus_amount),
                "Gaji Kotor": ExcelExportService._decimal_to_float(ps.gross_salary),
                "BPJS KES": ExcelExportService._decimal_to_float(ps.bpjs_kes_employee),
                "BPJS JHT": ExcelExportService._decimal_to_float(ps.bpjs_jht_employee),
                "BPJS JP": ExcelExportService._decimal_to_float(ps.bpjs_jp_employee),
                "PPh 21": ExcelExportService._decimal_to_float(ps.pph21_tax),
                "Potongan Lain": ExcelExportService._decimal_to_float(ps.other_deductions),
                "Kasbon": ExcelExportService._decimal_to_float(ps.kasbon_deduction),
                "Gaji Bersih": ExcelExportService._decimal_to_float(ps.net_salary),
            })

        df = pd.DataFrame(rows)
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Payslips")
            ws = writer.sheets["Payslips"]
            ExcelExportService._freeze_header(ws)
            ExcelExportService._auto_width(ws)
            # Currency columns: 4-15 (Gaji Pokok through Gaji Bersih)
            ExcelExportService._apply_currency_format(ws, list(range(4, 16)))

        return output.getvalue()

    @staticmethod
    def export_payroll_summary(payroll_run_id: int, db: Session) -> bytes:
        """Export payroll summary grouped by department."""
        payslips = (
            db.query(Payslip)
            .filter(Payslip.payroll_run_id == payroll_run_id)
            .all()
        )

        employee_ids = [p.employee_id for p in payslips]
        employees = {
            e.id: e
            for e in db.query(Employee).filter(Employee.id.in_(employee_ids)).all()
        } if employee_ids else {}

        dept_ids = list(set(e.department_id for e in employees.values() if e.department_id))
        departments = {
            d.id: d.name
            for d in db.query(Department).filter(Department.id.in_(dept_ids)).all()
        } if dept_ids else {}

        # Group by department
        dept_summary = {}
        for ps in payslips:
            emp = employees.get(ps.employee_id)
            dept_name = departments.get(emp.department_id, "Tanpa Departemen") if emp and emp.department_id else "Tanpa Departemen"
            if dept_name not in dept_summary:
                dept_summary[dept_name] = {
                    "count": 0,
                    "gross": 0.0,
                    "deductions": 0.0,
                    "tax": 0.0,
                    "net": 0.0,
                }
            dept_summary[dept_name]["count"] += 1
            dept_summary[dept_name]["gross"] += ExcelExportService._decimal_to_float(ps.gross_salary)
            dept_summary[dept_name]["deductions"] += (
                ExcelExportService._decimal_to_float(ps.bpjs_kes_employee)
                + ExcelExportService._decimal_to_float(ps.bpjs_jht_employee)
                + ExcelExportService._decimal_to_float(ps.bpjs_jp_employee)
                + ExcelExportService._decimal_to_float(ps.other_deductions)
                + ExcelExportService._decimal_to_float(ps.kasbon_deduction)
            )
            dept_summary[dept_name]["tax"] += ExcelExportService._decimal_to_float(ps.pph21_tax)
            dept_summary[dept_name]["net"] += ExcelExportService._decimal_to_float(ps.net_salary)

        rows = []
        grand_total = {"count": 0, "gross": 0.0, "deductions": 0.0, "tax": 0.0, "net": 0.0}
        for dept_name, summary in dept_summary.items():
            rows.append({
                "Departemen": dept_name,
                "Jumlah Karyawan": summary["count"],
                "Total Gaji Kotor": summary["gross"],
                "Total Potongan": summary["deductions"],
                "Total Pajak": summary["tax"],
                "Total Gaji Bersih": summary["net"],
            })
            for key in grand_total:
                grand_total[key] += summary[key]

        # Grand total row
        rows.append({
            "Departemen": "GRAND TOTAL",
            "Jumlah Karyawan": grand_total["count"],
            "Total Gaji Kotor": grand_total["gross"],
            "Total Potongan": grand_total["deductions"],
            "Total Pajak": grand_total["tax"],
            "Total Gaji Bersih": grand_total["net"],
        })

        df = pd.DataFrame(rows)
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Payroll Summary")
            ws = writer.sheets["Payroll Summary"]
            ExcelExportService._freeze_header(ws)
            ExcelExportService._auto_width(ws)
            ExcelExportService._apply_currency_format(ws, [3, 4, 5, 6])

        return output.getvalue()

    @staticmethod
    def export_bpjs_recap(company_id: int, period_month: int, period_year: int, db: Session) -> bytes:
        """Export BPJS recap for a specific month/year."""
        # Find payroll run for the period
        period_str = f"{period_year}-{period_month:02d}"
        payroll_run = (
            db.query(PayrollRun)
            .filter(
                PayrollRun.company_id == company_id,
                PayrollRun.payroll_period.like(f"{period_str}%"),
            )
            .first()
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "BPJS Recap"
        headers = ["Kode", "Nama", "No BPJS KES", "BPJS KES (Emp)", "BPJS JHT (Emp)", "BPJS JP (Emp)", "Total Karyawan"]
        ws.append(headers)

        if payroll_run:
            payslips = (
                db.query(Payslip)
                .filter(Payslip.payroll_run_id == payroll_run.id)
                .all()
            )
            employee_ids = [p.employee_id for p in payslips]
            employees = {
                e.id: e
                for e in db.query(Employee).filter(Employee.id.in_(employee_ids)).all()
            } if employee_ids else {}

            for ps in payslips:
                emp = employees.get(ps.employee_id)
                if not emp:
                    continue
                bpjs_kes = ExcelExportService._decimal_to_float(ps.bpjs_kes_employee)
                bpjs_jht = ExcelExportService._decimal_to_float(ps.bpjs_jht_employee)
                bpjs_jp = ExcelExportService._decimal_to_float(ps.bpjs_jp_employee)
                total = bpjs_kes + bpjs_jht + bpjs_jp
                ws.append([
                    emp.employee_code,
                    emp.full_name,
                    emp.bpjs_kesehatan_number or "",
                    bpjs_kes,
                    bpjs_jht,
                    bpjs_jp,
                    total,
                ])

        ExcelExportService._freeze_header(ws)
        ExcelExportService._auto_width(ws)
        ExcelExportService._apply_currency_format(ws, [4, 5, 6, 7])

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def generate_attendance_template() -> bytes:
        """Generate an empty attendance import template with a sample row."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Template"
        headers = [
            "employee_code",
            "attendance_date",
            "status",
            "check_in_time",
            "check_out_time",
            "hours_worked",
            "is_late",
            "late_minutes",
            "notes",
        ]
        ws.append(headers)
        ws.append([
            "EMP0001",
            "2026-06-26",
            "PRESENT",
            "08:00:00",
            "17:00:00",
            8,
            "FALSE",
            0,
            "",
        ])
        ws.append([
            "EMP0001",
            "2026-06-27",
            "SICK",
            "",
            "",
            "",
            "FALSE",
            0,
            "Sakit demam",
        ])

        ExcelExportService._freeze_header(ws)
        ExcelExportService._auto_width(ws)

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_attendance(company_id: int, period_month: int, period_year: int, db: Session) -> bytes:
        """Export attendance records for a specific company and month/year."""
        start_date = date(period_year, period_month, 1)
        if period_month == 12:
            end_date = date(period_year + 1, 1, 1)
        else:
            end_date = date(period_year, period_month + 1, 1)

        records = (
            db.query(AttendanceRecord)
            .join(Employee, AttendanceRecord.employee_id == Employee.id)
            .filter(
                Employee.company_id == company_id,
                AttendanceRecord.attendance_date >= start_date,
                AttendanceRecord.attendance_date < end_date,
            )
            .order_by(AttendanceRecord.attendance_date.asc(), Employee.employee_code.asc())
            .all()
        )

        employee_ids = [r.employee_id for r in records]
        employees = {
            e.id: e
            for e in db.query(Employee).filter(Employee.id.in_(employee_ids)).all()
        } if employee_ids else {}

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        headers = [
            "employee_code",
            "full_name",
            "attendance_date",
            "status",
            "check_in_time",
            "check_out_time",
            "hours_worked",
            "is_late",
            "late_minutes",
            "notes",
        ]
        ws.append(headers)

        for rec in records:
            emp = employees.get(rec.employee_id)
            ws.append([
                emp.employee_code if emp else "",
                emp.full_name if emp else "",
                rec.attendance_date.isoformat(),
                rec.status,
                rec.check_in_time or "",
                rec.check_out_time or "",
                ExcelExportService._decimal_to_float(rec.hours_worked),
                "TRUE" if rec.is_late else "FALSE",
                rec.late_minutes or 0,
                rec.notes or "",
            ])

        ExcelExportService._freeze_header(ws)
        ExcelExportService._auto_width(ws)

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_tax_recap(company_id: int, period_month: int, period_year: int, db: Session) -> bytes:
        """Export tax (PPh 21) recap for a specific month/year."""
        period_str = f"{period_year}-{period_month:02d}"
        payroll_run = (
            db.query(PayrollRun)
            .filter(
                PayrollRun.company_id == company_id,
                PayrollRun.payroll_period.like(f"{period_str}%"),
            )
            .first()
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Tax Recap"
        headers = ["Kode", "Nama", "NPWP", "Status PTKP", "Penghasilan Bruto", "PPh 21"]
        ws.append(headers)

        if payroll_run:
            payslips = (
                db.query(Payslip)
                .filter(Payslip.payroll_run_id == payroll_run.id)
                .all()
            )
            employee_ids = [p.employee_id for p in payslips]
            employees = {
                e.id: e
                for e in db.query(Employee).filter(Employee.id.in_(employee_ids)).all()
            } if employee_ids else {}

            for ps in payslips:
                emp = employees.get(ps.employee_id)
                if not emp:
                    continue
                ws.append([
                    emp.employee_code,
                    emp.full_name,
                    emp.npwp or "",
                    emp.ptkp_status or "",
                    ExcelExportService._decimal_to_float(ps.gross_salary),
                    ExcelExportService._decimal_to_float(ps.pph21_tax),
                ])

        ExcelExportService._freeze_header(ws)
        ExcelExportService._auto_width(ws)
        ExcelExportService._apply_currency_format(ws, [5, 6])

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    # ─── Overtime ───────────────────────────────────────────────────────────────

    @staticmethod
    def export_overtime(company_id: int, period_month: int, period_year: int, db: Session) -> bytes:
        """Export overtime records for a specific company and month/year."""
        start_date = date(period_year, period_month, 1)
        if period_month == 12:
            end_date = date(period_year + 1, 1, 1)
        else:
            end_date = date(period_year, period_month + 1, 1)

        records = (
            db.query(OvertimeRecord)
            .join(Employee, OvertimeRecord.employee_id == Employee.id)
            .filter(
                Employee.company_id == company_id,
                OvertimeRecord.overtime_date >= start_date,
                OvertimeRecord.overtime_date < end_date,
            )
            .order_by(OvertimeRecord.overtime_date.asc(), Employee.employee_code.asc())
            .all()
        )

        employee_ids = [r.employee_id for r in records]
        employees = {
            e.id: e
            for e in db.query(Employee).filter(Employee.id.in_(employee_ids)).all()
        } if employee_ids else {}

        wb = Workbook()
        ws = wb.active
        ws.title = "Overtime"
        headers = [
            "employee_code",
            "full_name",
            "overtime_date",
            "overtime_type",
            "hours",
            "multiplier",
            "approval_status",
            "notes",
        ]
        ws.append(headers)

        for rec in records:
            emp = employees.get(rec.employee_id)
            ws.append([
                emp.employee_code if emp else "",
                emp.full_name if emp else "",
                rec.overtime_date.isoformat(),
                rec.overtime_type,
                ExcelExportService._decimal_to_float(rec.hours),
                ExcelExportService._decimal_to_float(rec.multiplier),
                rec.approval_status,
                rec.notes or "",
            ])

        ExcelExportService._freeze_header(ws)
        ExcelExportService._auto_width(ws)

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def generate_overtime_template() -> bytes:
        """Generate an empty overtime import template with sample rows."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Overtime Template"
        headers = [
            "employee_code",
            "overtime_date",
            "overtime_type",
            "hours",
            "notes",
        ]
        ws.append(headers)
        ws.append(["EMP0001", "2026-06-26", "WEEKDAY", 3, "Lembur proyek A"])
        ws.append(["EMP0001", "2026-06-27", "WEEKEND", 4, "Lembur hari Sabtu"])
        ws.append(["EMP0001", "2026-06-17", "HOLIDAY", 2, "Lembur hari libur nasional"])

        ExcelExportService._freeze_header(ws)
        ExcelExportService._auto_width(ws)

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    # ─── Bonus ─────────────────────────────────────────────────────────────────

    @staticmethod
    def export_bonuses(company_id: int, db: Session) -> bytes:
        """Export bonus records for a company as Excel."""
        rows = db.query(Bonus).join(Employee, Bonus.employee_id == Employee.id).filter(
            Employee.company_id == company_id
        ).order_by(Bonus.bonus_date.desc()).all()

        employee_ids = [r.employee_id for r in rows]
        employees = {
            e.id: e
            for e in db.query(Employee).filter(Employee.id.in_(employee_ids)).all()
        } if employee_ids else {}

        type_ids = [r.bonus_type_id for r in rows]
        types = {
            t.id: t
            for t in db.query(BonusType).filter(BonusType.id.in_(type_ids)).all()
        } if type_ids else {}

        wb = Workbook()
        ws = wb.active
        ws.title = "Bonuses"
        headers = ["employee_code", "bonus_type_code", "amount", "bonus_date", "description", "approval_status"]
        ws.append(headers)

        for r in rows:
            emp = employees.get(r.employee_id)
            bt = types.get(r.bonus_type_id)
            ws.append([
                emp.employee_code if emp else "",
                bt.code if bt else "",
                ExcelExportService._decimal_to_float(r.amount),
                r.bonus_date.isoformat(),
                r.description or "",
                r.approval_status,
            ])

        ExcelExportService._freeze_header(ws)
        ExcelExportService._auto_width(ws)
        ExcelExportService._apply_currency_format(ws, [3])

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def generate_bonus_template(company_id: int, db: Session) -> bytes:
        """Generate an empty bonus import template with available type codes."""
        types = db.query(BonusType).filter(
            BonusType.company_id == company_id,
            BonusType.is_active == True,
        ).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Bonus Template"
        headers = ["employee_code", "bonus_type_code", "amount", "bonus_date", "description"]
        ws.append(headers)
        ws.append(["EMP0001", types[0].code if types else "BONUS001", 1000000, "2026-01-15", "Bonus kinerja"])

        ExcelExportService._freeze_header(ws)
        ExcelExportService._auto_width(ws)
        ExcelExportService._apply_currency_format(ws, [3])

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    # ─── THR ───────────────────────────────────────────────────────────────────

    @staticmethod
    def export_thr(company_id: int, db: Session) -> bytes:
        """Export THR records for a company as Excel."""
        rows = db.query(THRRecord).filter(THRRecord.company_id == company_id).order_by(
            THRRecord.thr_year.desc(), THRRecord.thr_date.desc()
        ).all()

        employee_ids = [r.employee_id for r in rows]
        employees = {
            e.id: e
            for e in db.query(Employee).filter(Employee.id.in_(employee_ids)).all()
        } if employee_ids else {}

        wb = Workbook()
        ws = wb.active
        ws.title = "THR"
        headers = ["employee_code", "thr_year", "religious_holiday", "amount", "thr_date", "calculation_basis", "description"]
        ws.append(headers)

        for r in rows:
            emp = employees.get(r.employee_id)
            ws.append([
                emp.employee_code if emp else "",
                r.thr_year,
                r.religious_holiday,
                ExcelExportService._decimal_to_float(r.amount),
                r.thr_date.isoformat(),
                r.calculation_basis,
                r.description or "",
            ])

        ExcelExportService._freeze_header(ws)
        ExcelExportService._auto_width(ws)
        ExcelExportService._apply_currency_format(ws, [4])

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def generate_thr_template() -> bytes:
        """Generate an empty THR import template."""
        wb = Workbook()
        ws = wb.active
        ws.title = "THR Template"
        headers = ["employee_code", "thr_year", "religious_holiday", "amount", "thr_date", "calculation_basis", "description"]
        ws.append(headers)
        ws.append(["EMP0001", 2026, "IDUL_FITRI", 5000000, "2026-04-01", "BASE_SALARY", "THR tahunan"])

        ExcelExportService._freeze_header(ws)
        ExcelExportService._auto_width(ws)
        ExcelExportService._apply_currency_format(ws, [4])

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    # ─── Reimbursement ─────────────────────────────────────────────────────────

    @staticmethod
    def export_reimbursements(company_id: int, db: Session) -> bytes:
        """Export reimbursement claims for a company as Excel."""
        rows = db.query(Reimbursement).join(Employee, Reimbursement.employee_id == Employee.id).filter(
            Employee.company_id == company_id
        ).order_by(Reimbursement.claim_date.desc()).all()

        employee_ids = [r.employee_id for r in rows]
        employees = {
            e.id: e
            for e in db.query(Employee).filter(Employee.id.in_(employee_ids)).all()
        } if employee_ids else {}

        type_ids = [r.reimbursement_type_id for r in rows]
        types = {
            t.id: t
            for t in db.query(ReimbursementType).filter(ReimbursementType.id.in_(type_ids)).all()
        } if type_ids else {}

        wb = Workbook()
        ws = wb.active
        ws.title = "Reimbursements"
        headers = ["employee_code", "reimbursement_type_code", "claim_amount", "approved_amount", "claim_date", "expense_date", "description", "approval_status"]
        ws.append(headers)

        for r in rows:
            emp = employees.get(r.employee_id)
            rt = types.get(r.reimbursement_type_id)
            ws.append([
                emp.employee_code if emp else "",
                rt.code if rt else "",
                ExcelExportService._decimal_to_float(r.claim_amount),
                ExcelExportService._decimal_to_float(r.approved_amount) if r.approved_amount else "",
                r.claim_date.isoformat(),
                r.expense_date.isoformat(),
                r.description or "",
                r.approval_status,
            ])

        ExcelExportService._freeze_header(ws)
        ExcelExportService._auto_width(ws)
        ExcelExportService._apply_currency_format(ws, [3, 4])

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def generate_reimbursement_template(company_id: int, db: Session) -> bytes:
        """Generate an empty reimbursement import template with available type codes."""
        types = db.query(ReimbursementType).filter(
            ReimbursementType.company_id == company_id,
            ReimbursementType.is_active == True,
        ).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Reimbursement Template"
        headers = ["employee_code", "reimbursement_type_code", "claim_amount", "claim_date", "expense_date", "description"]
        ws.append(headers)
        ws.append(["EMP0001", types[0].code if types else "REIMB001", 500000, "2026-01-15", "2026-01-10", "Transportasi dinas"])

        ExcelExportService._freeze_header(ws)
        ExcelExportService._auto_width(ws)
        ExcelExportService._apply_currency_format(ws, [3])

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    # ─── Kasbon ────────────────────────────────────────────────────────────────

    @staticmethod
    def export_kasbon(company_id: int, db: Session) -> bytes:
        """Export kasbon/loan records for a company as Excel."""
        rows = db.query(KasbonRequest).join(Employee, KasbonRequest.employee_id == Employee.id).filter(
            Employee.company_id == company_id
        ).order_by(KasbonRequest.request_date.desc()).all()

        employee_ids = [r.employee_id for r in rows]
        employees = {
            e.id: e
            for e in db.query(Employee).filter(Employee.id.in_(employee_ids)).all()
        } if employee_ids else {}

        wb = Workbook()
        ws = wb.active
        ws.title = "Kasbon"
        headers = ["employee_code", "kasbon_number", "principal_amount", "interest_rate", "number_of_installments", "request_date", "purpose", "status"]
        ws.append(headers)

        for r in rows:
            emp = employees.get(r.employee_id)
            ws.append([
                emp.employee_code if emp else "",
                r.kasbon_number,
                ExcelExportService._decimal_to_float(r.principal_amount),
                ExcelExportService._decimal_to_float(r.interest_rate) if r.interest_rate else 0,
                r.number_of_installments,
                r.request_date.isoformat(),
                r.purpose,
                r.status,
            ])

        ExcelExportService._freeze_header(ws)
        ExcelExportService._auto_width(ws)
        ExcelExportService._apply_currency_format(ws, [3])

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def generate_kasbon_template() -> bytes:
        """Generate an empty kasbon import template."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Kasbon Template"
        headers = ["employee_code", "kasbon_number", "principal_amount", "interest_rate", "number_of_installments", "request_date", "purpose"]
        ws.append(headers)
        ws.append(["EMP0001", "KSB-202601-0001", 10000000, 0, 6, "2026-01-15", "Kebutuhan darurat"])

        ExcelExportService._freeze_header(ws)
        ExcelExportService._auto_width(ws)
        ExcelExportService._apply_currency_format(ws, [3])

        output = BytesIO()
        wb.save(output)
        return output.getvalue()
